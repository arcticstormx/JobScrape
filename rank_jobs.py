import os
import re
from pathlib import Path
from typing import List, Tuple, Optional, Dict

import pandas as pd
from openpyxl.styles import PatternFill


def read_text_file(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="ignore")


def read_pdf_text(path: Path) -> str:
    try:
        from pypdf import PdfReader  # type: ignore
    except Exception as e:
        raise RuntimeError(
            "pypdf is required to read PDF resume. Please add it to requirements.txt"
        ) from e
    reader = PdfReader(str(path))
    parts: List[str] = []
    for page in reader.pages:
        try:
            parts.append(page.extract_text() or "")
        except Exception:
            continue
    return "\n".join(parts)


def load_resume_text(resume_path: Path) -> Tuple[str, str]:
    """Deprecated: resume usage removed."""
    return "", ""


def normalize_text(s: str) -> str:
    s = s.replace("\r", "\n")
    s = "\n".join(line.strip() for line in s.splitlines())
    # collapse multiple newlines/spaces
    s = " ".join(s.split())
    return s.strip()


def load_rank_instructions() -> Optional[str]:
    """Deprecated: similarity ranking removed; instructions no longer used."""
    return None


def build_resume_query(resume_text: str, instructions: Optional[str]) -> str:
    """Deprecated: similarity ranking removed."""
    return ""


# -------------------------
# Rule-based Rubric Scoring (Ranking AI)
# -------------------------

TOP_PROP_FIRMS = [
    "Citadel",
    "Jane Street",
    "Optiver",
    "IMC",
    "DRW",
    "Hudson River Trading",
    "Jump",
    "Five Rings",
    "SIG",
    "Flow Traders",
    "Two Sigma Securities",
    "Tower",
    "Akuna",
    "Wolverine",
    "Belvedere",
    "CTC",
    "DV Trading",
    "Virtu",
    "XTX",
]

TITLE_KEYWORDS = [
    "assistant trader",
    "execution trader",
    "trading assistant",
    "trading operations",
    "trade support",
    "trading specialist",
    "investment associate",
    "broker dealer",
]

EQUITY_KEYWORDS = ["equities", "options", "etf", "stock markets", "brokerage"]

# Brokerage/Client-Facing Fit (+2)
BROKERAGE_CLIENT_KEYWORDS = [
    "brokerage",
    "broker-dealer",
    "broker dealer",
    "client service",
    "corporate actions",
    "trade support",
    "trading operations",
    "active trader",
]

TRADING_DESK_KEYWORDS = ["trading desk"]

LICENSE_KEYWORDS = ["series 7", "series 63"]

PREFERRED_LOCATIONS = [
    "new york",
    "nyc",
    "chicago",
    "boston",
    "washington dc",
    "d.c.",
    "dc",
]
FLORIDA_LOCATIONS = ["florida", "miami", "tampa", "orlando"]

SENIORITY_KEYWORDS = ["senior", "vp", "director", "principal", "lead"]


def score_job_posting(title: str, company: str, location: str, description: str):
    score = 0
    breakdown: List[str] = []

    title_lc = str(title or "").lower()
    company_lc = str(company or "").lower()
    location_lc = str(location or "").lower()
    desc_lc = str(description or "").lower()

    if any(firm.lower() in company_lc for firm in TOP_PROP_FIRMS):
        score += 4
        breakdown.append("+4 Top Prop Firm")

    if any(keyword in title_lc for keyword in TITLE_KEYWORDS):
        score += 3
        breakdown.append("+3 Title Match")

    if any(keyword in desc_lc for keyword in EQUITY_KEYWORDS):
        score += 2
        breakdown.append("+2 Equities/Markets Focus")

    if any(keyword in desc_lc for keyword in TRADING_DESK_KEYWORDS):
        score += 2
        breakdown.append("+2 Trading Desk")

    # Brokerage/Client-Facing Fit (+2) — check title or description
    if any(kw in title_lc or kw in desc_lc for kw in BROKERAGE_CLIENT_KEYWORDS):
        score += 2
        breakdown.append("+2 Brokerage/Client-Facing Fit")

    if any(keyword in desc_lc for keyword in LICENSE_KEYWORDS):
        score += 1
        breakdown.append("+1 Series License")

    if any(loc in location_lc for loc in PREFERRED_LOCATIONS):
        score += 1
        breakdown.append("+1 Preferred Location")

    if any(keyword in title_lc for keyword in SENIORITY_KEYWORDS):
        score -= 5
        breakdown.append("-5 Seniority Penalty")

    if any(loc in location_lc for loc in FLORIDA_LOCATIONS):
        score -= 1
        breakdown.append("-1 Florida Penalty")

    # New threshold: Top Pick at >= 10 (max now 15)
    top_pick = score >= 10
    return score, breakdown, top_pick


def build_ranking_ai_sheet(all_df: pd.DataFrame) -> pd.DataFrame:
    """Return a DataFrame scored by the rubric, sorted by ai_score desc.

    Adds columns: ai_score (int), ai_top_pick (bool), ai_breakdown (str)
    """
    if all_df.empty:
        return all_df.copy()

    # Remove postings that would incur a seniority penalty entirely
    if "title" in all_df.columns:
        title_lc = all_df["title"].astype(str).str.lower()
        seniority_pat = "|".join(map(re.escape, SENIORITY_KEYWORDS))
        all_df = all_df[~title_lc.str.contains(seniority_pat, na=False)]

    rows = []
    for _, r in all_df.iterrows():
        title = r.get("title", "")
        company = r.get("company", "")
        location = r.get("location", "")
        desc = r.get("description", "")
        score, breakdown, top_pick = score_job_posting(title, company, location, desc)
        rows.append((score, top_pick, "; ".join(breakdown)))

    out = all_df.copy()
    scores, top_picks, breakdowns = zip(*rows) if rows else ([], [], [])
    out.insert(0, "ai_score", list(scores))
    out.insert(1, "ai_top_pick", list(top_picks))
    out.insert(2, "ai_breakdown", list(breakdowns))
    out = out.sort_values(by=["ai_score"], ascending=False)
    return out


def _highlight_top_picks(writer, sheet_name: str, df: pd.DataFrame) -> None:
    """Highlight rows where ai_top_pick is True in light green.

    Applies styling directly to the openpyxl worksheet via the pandas ExcelWriter.
    """
    if "ai_top_pick" not in df.columns:
        return
    try:
        ws = writer.sheets[sheet_name]
    except Exception:
        return
    fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ncols = len(df.columns)
    # Iterate DataFrame rows; header is row 1, data starts at row 2
    for i, is_top in enumerate(df["ai_top_pick"].tolist(), start=2):
        if bool(is_top):
            for col_idx in range(1, ncols + 1):
                ws.cell(row=i, column=col_idx).fill = fill

def get_openai_client():
    return None


def embed_texts(client, texts: List[str], model: str) -> List[List[float]]:
    raise RuntimeError("Embeddings not supported: similarity ranking removed.")


def rank_with_openai(resume_text: str, texts: List[str]) -> Tuple[List[float], str]:
    raise RuntimeError("OpenAI disabled: similarity ranking removed.")


def rank_with_local(resume_text: str, texts: List[str]) -> Tuple[List[float], str]:
    raise RuntimeError("Local similarity disabled: ranking uses rubric only.")


def cosine_sim(a, b) -> float:
    return 0.0


def main():
    # Inputs
    jobs_excel = Path(os.getenv("JOBS_EXCEL", "Jobs/all_jobs.xlsx"))
    out_excel = Path(os.getenv("OUT_EXCEL", "Jobs/all_jobs_ranked.xlsx"))

    if not jobs_excel.exists():
        raise FileNotFoundError(f"Jobs Excel not found at {jobs_excel}")

    # Load jobs workbook fully into memory to allow safe deletion later
    sheets_map: Dict[str, pd.DataFrame] = pd.read_excel(jobs_excel, sheet_name=None)
    sheet_order = list(sheets_map.keys())
    if not sheet_order:
        with pd.ExcelWriter(out_excel) as writer:
            pd.DataFrame().to_excel(writer, sheet_name="Empty", index=False)
        print("No sheets found. Wrote Empty sheet.")
        return

    # Helper to convert a row to text for embedding
    def row_to_text(row: pd.Series) -> str:
        title = str(row.get("title", ""))
        company = str(row.get("company", ""))
        location = str(row.get("location", ""))
        desc = str(row.get("description", ""))
        if len(desc) > 4000:
            desc = desc[:4000]
        return normalize_text(
            f"Title: {title}\nCompany: {company}\nLocation: {location}\nDescription: {desc}"
        )

    # Function to rank a single DataFrame
    def rank_df(df: pd.DataFrame) -> Tuple[pd.DataFrame, str]:
        if df.empty:
            return df.copy(), "no_data"
        # Remove postings that would incur a seniority penalty entirely
        if "title" in df.columns:
            title_lc = df["title"].astype(str).str.lower()
            seniority_pat = "|".join(map(re.escape, SENIORITY_KEYWORDS))
            df = df[~title_lc.str.contains(seniority_pat, na=False)]
        # Rank using rubric scoring
        out = build_ranking_ai_sheet(df)
        return out, "rubric"

    # Build an All dataframe from input (prefer the input 'All' if present)
    if "All" in sheet_order:
        input_all_df = sheets_map["All"]
    else:
        input_all_df = pd.concat([sheets_map[s] for s in sheet_order], ignore_index=True)

    # Rank each sheet independently and write to output with same sheet names
    methods_used: Dict[str, str] = {}
    with pd.ExcelWriter(out_excel) as writer:
        # First, add the rule-based 'Ranking AI' sheet based on all jobs
        ranking_ai_df = build_ranking_ai_sheet(input_all_df)
        ranking_ai_df.to_excel(writer, sheet_name="Ranking AI", index=False)
        methods_used["Ranking AI"] = "rubric"

        # Ensure 'All' (if exists) is written first to mimic original ordering
        ordered = (
            ["All"] + [s for s in sheet_order if s != "All"]
            if "All" in sheet_order
            else sheet_order
        )
        for name in ordered:
            df_sheet = sheets_map[name]
            ranked_df, method = rank_df(df_sheet)
            methods_used[name] = method
            ranked_df.to_excel(writer, sheet_name=name, index=False)

    # Report summary
    method_summary = ", ".join(f"{k}:{v}" for k, v in methods_used.items())
    print(
        f"Ranked workbook '{jobs_excel.name}' with methods per sheet [{method_summary}] (resume not used).\n"
        f"Wrote {out_excel}"
    )

    # Remove the unranked source workbook so only the ranked file remains
    keep_unranked = os.getenv("KEEP_UNRANKED", "").lower() in {"1", "true", "yes", "y"}
    try:
        if not keep_unranked and jobs_excel.exists() and jobs_excel.resolve() != out_excel.resolve():
            jobs_excel.unlink(missing_ok=True)
            print(f"Removed source workbook: {jobs_excel}")
    except Exception as e:
        print(f"Warning: could not remove source workbook {jobs_excel}: {e}")


if __name__ == "__main__":
    main()
